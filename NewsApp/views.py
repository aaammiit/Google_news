from django.shortcuts import render,redirect
import feedparser
import nltk
# from nltk.tokenize import word_tokenize
import threading
import schedule
import time
import tenacity
import smtplib
from dateutil import parser
from bs4 import BeautifulSoup
from urllib3 import request
import requests
import pandas as pd
from datetime import datetime,timedelta
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
import os
import json
from django.contrib import messages
from .models import *

SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 587
SMTP_USERNAME = 'amistreetecom0101@gmail.com'
SMTP_PASSWORD = 'xmxj ztoo dfvw bxtm'
sender_email = 'amistreetecom0101@gmail.com'
Receiver_email='newstracker@krimasolutions.com'

subject = "News Alert"

def is_within_last_5_days(date):
    today = datetime.today()
    five_days_ago = today - timedelta(days=2)
    return five_days_ago.date() <= date <= today.date()

def serialize_datetime(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    raise TypeError("Type not serializable")

dt = datetime.now()
json_data = json.dumps(dt, default=serialize_datetime)

Record=[]
def Home(request):
    # data=request.session.get('data')
    # data={'data':data}
    # data=[]
    data=News_model.objects.all()
    data={'data':data}
    return render(request,'index.html',data)



@tenacity.retry(wait=tenacity.wait_exponential(min=4, max=10), stop=tenacity.stop_after_attempt(3))
def Search_News(request):
    News=[]
    keyurl=''
    if request.method=='POST':
        url=request.POST.get('url')
        time_period=request.POST.get('time_period')
        keyurl=url

        time_period=time_period
        if url != 'nome' and time_period !='none':
            base_url=f'https://news.google.com/search?q={url}20when{time_period}&hl=en-IN&gl=IN&ceid=IN%3Aen'
            # print(base_url)
            response = requests.get(base_url)
            soup = BeautifulSoup(response.content, 'html.parser')
            new_records=[]
            for article in soup.find_all('article'):
                title_tag = article.find_all('a')[1]
                if title_tag:
                    title = title_tag.get_text(strip=True)
                    relative_url = title_tag.get('href')
                    if relative_url and relative_url.startswith('./'):
                        url = 'https://news.google.com' + relative_url[1:]
                        urls=url
                    else:
                        urls=relative_url    
                    date_tag = article.find('time')
                    
                    if date_tag:
                        date = date_tag.get('datetime')
                        date = datetime.strptime(date, '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
                        date=datetime.strptime(date, '%Y-%m-%d').date()
                        date_str = date.isoformat()
                        time_ago=date_tag.text       
                record={
                            'tilte':title,
                            'link':urls,
                            'date':date_str,
                            'ago':time_ago
                        } 
                if record not in Record:
                    new_records.append(record)
        News.extend(new_records)
    else:
        print('url and time period select None')
    new_list=[]
    sorted_list = sorted(News, key=lambda x: x['date'], reverse=True)
    News1 = sorted_list
    request.session['data']=News1

    # for i in sorted_list:
    #     if i not in new_list:
    #         print(len(new_list))

    print(len(News1))
    if len(News1) !=0:
       data={'data':News1,'url':keyurl}
       return render(request,'news.html',data)
    else:
        msg={'msg':'Sorry, not any news to fetch.'}
        return render(request,'news.html',msg)
        

def Clear_news(request):
    request.session.clear()
    return redirect('/')


@tenacity.retry(wait=tenacity.wait_exponential(min=4, max=10), stop=tenacity.stop_after_attempt(3))
def Send_file(request):
    url01='https://news.google.com/search?q=chief%20compliance%20officer%20appoint%20when%3A7d&hl=en-IN&gl=IN&ceid=IN%3Aen'
    url02='https://news.google.com/search?q=general%20counsel%20appoint%20when%3A7d&hl=en-IN&gl=IN&ceid=IN%3Aen'
    url03='https://news.google.com/search?q=chief%20risk%20officer%20appoint%20when%3A7d&hl=en-IN&gl=IN&ceid=IN%3Aen'
    url04='https://news.google.com/search?q=chief%20legal%20officer%20appoint%20when%3A7d&hl=en-IN&gl=IN&ceid=IN%3Aen'
    url05='https://news.google.com/search?q=chief%20sustainability%20officer%20appoint%20when%3A7d&hl=en-IN&gl=IN&ceid=IN%3Aen'
    url06='https://news.google.com/search?q=integrity%20officer%20appoint%20when%3A7d&hl=en-IN&gl=IN&ceid=IN%3Aen'
    url07='https://news.google.com/search?q=ethics%20officer%20appoint%20when%3A7d&hl=en-IN&gl=IN&ceid=IN%3Aen'
    url08='https://news.google.com/search?q=chief%20diversity%C2%A0officer%C2%A0appoint%20when%3A7d&hl=en-IN&gl=IN&ceid=IN%3Aen'

    url_list=[url01,url02,url03,url04,url05,url06,url07,url08]
    new_records=[]
    news=[]
    for url in range(len(url_list)):
        today = datetime.today() 
        response = requests.get(url_list[url])
        soup = BeautifulSoup(response.content, 'html.parser')
        for article in soup.find_all('article'):
            title_tag = article.find_all('a')[1]
            if title_tag:
                title = title_tag.get_text(strip=True)
                relative_url = title_tag.get('href')
                if relative_url and relative_url.startswith('./'):
                    url = 'https://news.google.com' + relative_url[1:]
                    urls=url
                else:
                    urls=relative_url    
            date_tag = article.find('time')
            if date_tag:
                date = date_tag.get('datetime')
                date = datetime.strptime(date, '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
                date=datetime.strptime(date, '%Y-%m-%d').date()
                date_str = date.isoformat()
                time_ago=date_tag.text 
                    
            record={
                    'tilte':title,
                    'link':urls,
                    'date':date_str,
                    'ago':time_ago
                }
            if record['tilte'] not in new_records:
                new_records.append(record)
            
        wb = openpyxl.Workbook()
        ws = wb.active
                
        ws['A1'] = 'Title'
        ws['B1'] = 'Link'
        ws['C1'] = 'Date'
        for i, record in enumerate(new_records):  
            ws[f'A{i+2}'] = record['tilte']
            ws[f'B{i+2}'] = record['link']
            ws[f'C{i+2}'] = record['date']
        date_str = datetime.now().strftime("%B%d")
        current_time_str = datetime.now().strftime("%H%M%S")
        filename = f'{date_str}News_records{current_time_str}.xlsx'
        wb.save(filename)

        subject = "News Alerts"
        body = "Here The Top And Latest News Tilte Url And Links"
        for record in new_records:
            body += f"Title: {record['tilte']}, Link: {record['link']}\n"

        msg = MIMEMultipart()
        msg['Subject'] = subject
        msg['From'] = sender_email
        msg['To'] = Receiver_email
        msg['Date'] = formatdate(localtime=True)

        try:
            msg.attach(MIMEText(body, 'plain'))

                    # Attach the Excel file
            attachment = open(f'{filename}', 'rb')
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f"attachment; filename= news_records.xlsx")
            msg.attach(part)

                    # Send the email
            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(sender_email, Receiver_email, msg.as_string())
            server.quit()
            print('sent')        
        except:
            print('err')
        print(len(new_records))
        new_records.clear()
    return redirect('/')

def Add_bucket(request):
    if request.method == 'POST':
        urlkeyword=request.POST.get('url')
        print(urlkeyword)
        try:
            news_ids = request.POST.getlist('news_ids')
            print(news_ids,type(news_ids))
            data = request.session.get('data')
            for id in news_ids:
                specific_data = data[int(id)]
                model_instance = News_model(keyword=urlkeyword,title=specific_data['tilte'], url=specific_data['link'], date=specific_data['date'], publised=specific_data['ago'])
                model_instance.save()
                print('saved')
            return redirect('/')
        except Exception as e:
            print(e)
            return redirect('/')
    else:
        return redirect('/')
        print('err')


def Remove(request,id):
    data=News_model.objects.filter(id=id)
    data.delete()
    return redirect('/')

def deleteAll(request):
    data=News_model.objects.all()
    try:
        data.delete()
        return redirect('/')
    except News_model.DoesNotExist as e:
        print(e)
        return redirect('/')


def Filter_news(request):
    filterKeyword1=''
    if request.method == 'POST':
        position = request.POST.get('position')
        if position:
            data=News_model.objects.filter(keyword__icontains=position)
            if position =='chief%20compliance%20officer%20appoint%':
                filterKeyword='Chief Compliance Officer Appoint'

            elif position =='general%20counsel%20appoint%':
                filterKeyword='General Counsel Appoint'

            elif position =='chief%20risk%20officer%20appoint%':
                filterKeyword='Chief Risk Officer Appoint'

            elif position =='chief%20legal%20officer%20appoint%':
                filterKeyword='Chief Legal Officer Appoint'

            elif position =='chief%20sustainability%20officer%20appoint%':
                filterKeyword='Chief Sustainability Officer Appoint'

            elif position =='integrity%20officer%20appoint%':
                filterKeyword='Integrity Officer Appoint'

            elif position =='ethics%20officer%20appoint%':
                filterKeyword='Ethics Officer Appoint'

            elif position =='chief%20diversity%20officer%20appoint%':
                filterKeyword='Chief Diversity Officer Appoint'
            filterKeyword1=filterKeyword
            data={'data':data,'keyword':filterKeyword1}
        return render(request,'index.html',data)  # Redirect after processing
    else:
        return redirect('/')


        
        
            